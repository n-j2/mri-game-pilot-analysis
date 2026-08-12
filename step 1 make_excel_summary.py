from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime, timezone
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =============================
# User settings
# =============================

DATA_DIR = Path("data")
OUTPUT_DIR = Path("analysis_output")
OUTPUT_EXCEL_FILE = OUTPUT_DIR / "participant_summary.xlsx"

N_MCQ_QUESTIONS = 7

# Meaningful completion definition from the study protocol
MEANINGFUL_MIN_NPCS = 8
MEANINGFUL_MIN_PRE_MCQ_ATTEMPTED = 5
MEANINGFUL_MIN_POST_MCQ_ATTEMPTED = 5


# =============================
# MCQ and NPC definitions
# =============================

QUESTION_IDS = [
    "q1",
    "q2",
    "q3",
    "q4",
    "q5",
    "q6",
    "q7",
]

QUESTION_LABELS = {
    "q1": "Q1",
    "q2": "Q2",
    "q3": "Q3",
    "q4": "Q4",
    "q5": "Q5",
    "q6": "Q6",
    "q7": "Q7",
}


# NPCs are listed in the approximate order in which they are encountered.
NPC_IDS = [
    "yoga_student_highfields",
    "jogging_student_highfields",
    "radiographer_1",
    "radiographer_2",
    "radiographer_3",
    "MRI_Participant",
    "chris_statistician",
    "nitish",
    "prof_dineen_offices",
    "emma_library",
    "journal_student_library",
    "dr_jawahar_imh",
]

NPC_LABELS = {
    "yoga_student_highfields": "Yoga Student",
    "jogging_student_highfields": "Jogging Student",
    "radiographer_1": "Entrance Radiographer",
    "radiographer_2": "Console Radiographer",
    "radiographer_3": "Contrast Radiographer",
    "MRI_Participant": "Research Participant",
    "chris_statistician": "Statistician",
    "nitish": "Postgraduate Student",
    "prof_dineen_offices": "Neuroimaging Professor",
    "emma_library": "Librarian",
    "journal_student_library": "Journalling Student",
    "dr_jawahar_imh": "Translational Clinician",
}


# =============================
# Time variables
# =============================

# Each seconds variable will also have a readable *_min_sec column.
TIME_COLUMNS_SEC = [
    "time_demographics_to_pre_mcq_sec",
    "time_gameplay_sec",
    "time_game_finish_to_post_mcq_sec",
    "time_post_mcq_to_post_questionnaire_sec",
]


# =============================
# Final output-column order
# =============================

# unique_npcs_interacted_with and npcs_interacted_with are inserted
# separately after time_game_finish_to_post_mcq_sec.
REQUESTED_COLUMNS = [
    "participant_id",
    "meaningful_completion",
    "meaningful_completion_failure_reason",

    "age_years",
    "sex",
    "gaming_frequency",

    "pre_mcq_attempted",
    "pre_mcq_questions_attempted",
    "pre_mcq_total_correct",
    "pre_mcq_questions_correct",

    "post_mcq_attempted",
    "post_mcq_questions_attempted",
    "post_mcq_total_correct",
    "post_mcq_questions_correct",

    "mcq_correct_change",

    "mri_anxiety_pre",
    "mri_anxiety_post",
    "mri_anxiety_change",

    "mri_willingness_pre",
    "mri_willingness_post",
    "mri_willingness_change",

    "technical_issues",
    "technical_issues_text",
    "most_useful_learning",
    "improvements",

    "aim_1",
    "aim_2",
    "aim_3",
    "aim_4",

    "time_demographics_to_pre_mcq_sec",
    "time_gameplay_sec",
    "time_game_finish_to_post_mcq_sec",
    "time_post_mcq_to_post_questionnaire_sec",
]


# =============================
# General helper functions
# =============================

def get_nested(
    data: dict[str, Any],
    path: list[str],
    default: Any = np.nan,
) -> Any:
    """
    Safely retrieve a value from a nested dictionary.

    Example:
        get_nested(data, ["pre_game", "mri_anxiety_pre"])
    """
    current: Any = data

    for key in path:
        if not isinstance(current, dict) or key not in current:
            return default

        current = current[key]

    if current is None:
        return default

    return current


def parse_iso_datetime(value: Any) -> datetime | None:
    """
    Parse an ISO timestamp such as:
        2026-06-04T10:01:50.806Z

    Returns a timezone-aware UTC datetime, or None if missing or invalid.
    """
    if value is None or not isinstance(value, str):
        return None

    text = value.strip()

    if text == "":
        return None

    if text.endswith("Z"):
        text = text[:-1] + "+00:00"

    try:
        parsed_datetime = datetime.fromisoformat(text)
    except ValueError:
        return None

    if parsed_datetime.tzinfo is None:
        parsed_datetime = parsed_datetime.replace(
            tzinfo=timezone.utc
        )

    return parsed_datetime.astimezone(
        timezone.utc
    )


def seconds_between(
    start: Any,
    end: Any,
) -> float:
    """
    Calculate duration in seconds between two ISO timestamps.

    Returns NaN if either timestamp is missing or invalid.
    Negative durations are treated as invalid.
    """
    start_datetime = parse_iso_datetime(start)
    end_datetime = parse_iso_datetime(end)

    if start_datetime is None or end_datetime is None:
        return np.nan

    duration_seconds = (
        end_datetime - start_datetime
    ).total_seconds()

    if duration_seconds < 0:
        return np.nan

    return round(
        duration_seconds,
        3,
    )


def format_seconds_as_min_sec(
    value: Any,
) -> str:
    """
    Convert seconds into a human-readable duration.

    Examples:
        45.2   -> 0m 45s
        191.3  -> 3m 11s
        3672   -> 1h 01m 12s
    """
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except TypeError:
        return ""

    try:
        total_seconds = int(
            round(float(value))
        )
    except (TypeError, ValueError):
        return ""

    if total_seconds < 0:
        return ""

    hours = total_seconds // 3600
    minutes = (
        total_seconds % 3600
    ) // 60
    seconds = total_seconds % 60

    if hours > 0:
        return (
            f"{hours}h "
            f"{minutes:02d}m "
            f"{seconds:02d}s"
        )

    return f"{minutes}m {seconds:02d}s"


def safe_int(
    value: Any,
) -> int | float:
    """
    Convert a value to an integer where possible.

    Returns NaN if missing or invalid.
    """
    if value is None:
        return np.nan

    try:
        if pd.isna(value):
            return np.nan
    except TypeError:
        pass

    try:
        return int(value)
    except (TypeError, ValueError):
        return np.nan


def numeric_change(
    post_value: Any,
    pre_value: Any,
) -> float:
    """
    Calculate post minus pre.

    Returns NaN if either value is missing or non-numeric.
    """
    try:
        if (
            pd.isna(pre_value)
            or pd.isna(post_value)
        ):
            return np.nan
    except TypeError:
        return np.nan

    try:
        return (
            float(post_value)
            - float(pre_value)
        )
    except (TypeError, ValueError):
        return np.nan


# =============================
# MCQ helper functions
# =============================

def get_mcq_answers(
    mcq_block: dict[str, Any],
) -> dict[str, Any]:
    """
    Safely retrieve the item-level MCQ answer dictionary.
    """
    answers = mcq_block.get(
        "answers",
        {},
    )

    if isinstance(answers, dict):
        return answers

    return {}


def get_attempted_question_ids(
    mcq_block: dict[str, Any],
) -> list[str]:
    """
    Return nominal question IDs marked answered == True.

    IDs are returned in nominal Q1-Q7 order, independently of the
    randomised display order used for that participant.
    """
    answers = get_mcq_answers(
        mcq_block
    )

    attempted_question_ids: list[str] = []

    for question_id in QUESTION_IDS:
        question_data = answers.get(
            question_id,
            {},
        )

        if (
            isinstance(question_data, dict)
            and question_data.get("answered") is True
        ):
            attempted_question_ids.append(
                question_id
            )

    return attempted_question_ids


def get_correct_question_ids(
    mcq_block: dict[str, Any],
) -> list[str]:
    """
    Return nominal question IDs marked correct == True.

    IDs are returned in nominal Q1-Q7 order.
    """
    answers = get_mcq_answers(
        mcq_block
    )

    correct_question_ids: list[str] = []

    for question_id in QUESTION_IDS:
        question_data = answers.get(
            question_id,
            {},
        )

        if (
            isinstance(question_data, dict)
            and question_data.get("correct") is True
        ):
            correct_question_ids.append(
                question_id
            )

    return correct_question_ids


def format_question_id_list(
    question_ids: list[str],
) -> str:
    """
    Convert nominal question IDs into a readable Excel string.

    Example:
        ["q1", "q3", "q7"] -> "Q1, Q3, Q7"
    """
    return ", ".join(
        QUESTION_LABELS[question_id]
        for question_id in question_ids
        if question_id in QUESTION_LABELS
    )


def count_attempted(
    mcq_block: dict[str, Any],
) -> int:
    """
    Count questions where answered == True.
    """
    return len(
        get_attempted_question_ids(
            mcq_block
        )
    )


def count_correct_from_answers(
    mcq_block: dict[str, Any],
) -> int:
    """
    Count questions where correct == True.

    Used if total_correct is missing.
    """
    return len(
        get_correct_question_ids(
            mcq_block
        )
    )


# =============================
# NPC helper functions
# =============================

def is_positive_conversation_count(
    value: Any,
) -> bool:
    """
    Return True when an NPC conversation-start count is numeric
    and greater than zero.
    """
    if value is None or isinstance(value, bool):
        return False

    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def get_interacted_npc_ids(
    npc_conversation_counts: Any,
) -> list[str]:
    """
    Return predefined NPC IDs with at least one recorded conversation.

    NPCs are returned in approximate game encounter order.
    """
    if not isinstance(
        npc_conversation_counts,
        dict,
    ):
        return []

    interacted_npc_ids: list[str] = []

    for npc_id in NPC_IDS:
        conversation_count = (
            npc_conversation_counts.get(
                npc_id,
                0,
            )
        )

        if is_positive_conversation_count(
            conversation_count
        ):
            interacted_npc_ids.append(
                npc_id
            )

    return interacted_npc_ids


def format_npc_list(
    npc_ids: list[str],
) -> str:
    """
    Convert internal NPC IDs into reader-facing NPC names.

    Conversation frequencies are deliberately not included.
    """
    return ", ".join(
        NPC_LABELS[npc_id]
        for npc_id in npc_ids
        if npc_id in NPC_LABELS
    )


# =============================
# Participant extraction
# =============================

def extract_participant_row(
    data: dict[str, Any],
) -> dict[str, Any]:
    """
    Extract the required analysis variables from one participant JSON file.
    """
    demographics = data.get(
        "demographics",
        {},
    )

    if not isinstance(demographics, dict):
        demographics = {}

    pre_mcq = data.get(
        "pre_mcq",
        {},
    )

    if not isinstance(pre_mcq, dict):
        pre_mcq = {}

    post_mcq = data.get(
        "post_mcq",
        {},
    )

    if not isinstance(post_mcq, dict):
        post_mcq = {}

    game = data.get(
        "game",
        {},
    )

    if not isinstance(game, dict):
        game = {}

    telemetry = game.get(
        "telemetry",
        {},
    )

    if not isinstance(telemetry, dict):
        telemetry = {}

    post_questionnaire = data.get(
        "post_questionnaire",
        {},
    )

    if not isinstance(
        post_questionnaire,
        dict,
    ):
        post_questionnaire = {}

    aim = post_questionnaire.get(
        "aim",
        {},
    )

    if not isinstance(aim, dict):
        aim = {}

    # -------------------------
    # MCQ completion and scores
    # -------------------------

    pre_attempted_question_ids = (
        get_attempted_question_ids(
            pre_mcq
        )
    )

    pre_correct_question_ids = (
        get_correct_question_ids(
            pre_mcq
        )
    )

    post_attempted_question_ids = (
        get_attempted_question_ids(
            post_mcq
        )
    )

    post_correct_question_ids = (
        get_correct_question_ids(
            post_mcq
        )
    )

    pre_mcq_attempted = len(
        pre_attempted_question_ids
    )

    post_mcq_attempted = len(
        post_attempted_question_ids
    )

    pre_mcq_questions_attempted = (
        format_question_id_list(
            pre_attempted_question_ids
        )
    )

    pre_mcq_questions_correct = (
        format_question_id_list(
            pre_correct_question_ids
        )
    )

    post_mcq_questions_attempted = (
        format_question_id_list(
            post_attempted_question_ids
        )
    )

    post_mcq_questions_correct = (
        format_question_id_list(
            post_correct_question_ids
        )
    )

    pre_mcq_total_correct = safe_int(
        pre_mcq.get(
            "total_correct",
            np.nan,
        )
    )

    if pd.isna(pre_mcq_total_correct):
        pre_mcq_total_correct = len(
            pre_correct_question_ids
        )

    post_mcq_total_correct = safe_int(
        post_mcq.get(
            "total_correct",
            np.nan,
        )
    )

    if pd.isna(post_mcq_total_correct):
        post_mcq_total_correct = len(
            post_correct_question_ids
        )

    mcq_correct_change = numeric_change(
        post_mcq_total_correct,
        pre_mcq_total_correct,
    )

    # -------------------------
    # Anxiety and willingness
    # -------------------------

    mri_anxiety_pre = get_nested(
        data,
        [
            "pre_game",
            "mri_anxiety_pre",
        ],
    )

    mri_anxiety_post = (
        post_questionnaire.get(
            "mri_anxiety_post",
            np.nan,
        )
    )

    mri_willingness_pre = get_nested(
        data,
        [
            "pre_game",
            "mri_willingness_pre",
        ],
    )

    mri_willingness_post = (
        post_questionnaire.get(
            "mri_willingness_post",
            np.nan,
        )
    )

    mri_anxiety_change = numeric_change(
        mri_anxiety_post,
        mri_anxiety_pre,
    )

    mri_willingness_change = numeric_change(
        mri_willingness_post,
        mri_willingness_pre,
    )

    # -------------------------
    # Gameplay telemetry
    # -------------------------

    npc_conversation_counts = telemetry.get(
        "npc_conversation_start_counts",
        {},
    )

    interacted_npc_ids = (
        get_interacted_npc_ids(
            npc_conversation_counts
        )
    )

    npcs_interacted_with = format_npc_list(
        interacted_npc_ids
    )

    unique_npcs_interacted_with = safe_int(
        telemetry.get(
            "unique_npcs_interacted_with",
            np.nan,
        )
    )

    # Retain the existing defensive fallback for the unique-NPC total.
    # It counts only predefined NPCs with a positive conversation count.
    if pd.isna(unique_npcs_interacted_with):
        unique_npcs_interacted_with = len(
            interacted_npc_ids
        )

    # Gameplay duration comes directly and exclusively from telemetry.
    time_gameplay_sec = telemetry.get(
        "time_played_seconds",
        np.nan,
    )

    # -------------------------
    # Meaningful completion
    # -------------------------

    meaningful_completion = (
        (
            not pd.isna(
                unique_npcs_interacted_with
            )
            and unique_npcs_interacted_with
            >= MEANINGFUL_MIN_NPCS
        )
        and pre_mcq_attempted
        >= MEANINGFUL_MIN_PRE_MCQ_ATTEMPTED
        and post_mcq_attempted
        >= MEANINGFUL_MIN_POST_MCQ_ATTEMPTED
    )

    failure_reasons: list[str] = []

    if (
        pd.isna(unique_npcs_interacted_with)
        or unique_npcs_interacted_with
        < MEANINGFUL_MIN_NPCS
    ):
        failure_reasons.append(
            f"NPCs < {MEANINGFUL_MIN_NPCS}"
        )

    if (
        pre_mcq_attempted
        < MEANINGFUL_MIN_PRE_MCQ_ATTEMPTED
    ):
        failure_reasons.append(
            "pre MCQ attempted < "
            f"{MEANINGFUL_MIN_PRE_MCQ_ATTEMPTED}"
        )

    if (
        post_mcq_attempted
        < MEANINGFUL_MIN_POST_MCQ_ATTEMPTED
    ):
        failure_reasons.append(
            "post MCQ attempted < "
            f"{MEANINGFUL_MIN_POST_MCQ_ATTEMPTED}"
        )

    meaningful_completion_failure_reason = (
        "; ".join(failure_reasons)
    )

    # -------------------------
    # Timestamps
    # -------------------------

    demographics_completed_at = data.get(
        "demographics_completed_at"
    )

    pre_mcq_completed_at = pre_mcq.get(
        "completed_at"
    )

    game_finished_at = game.get(
        "finished_at"
    )

    post_mcq_completed_at = post_mcq.get(
        "completed_at"
    )

    post_questionnaire_completed_at = (
        post_questionnaire.get(
            "completed_at"
        )
    )

    time_demographics_to_pre_mcq_sec = (
        seconds_between(
            demographics_completed_at,
            pre_mcq_completed_at,
        )
    )

    time_game_finish_to_post_mcq_sec = (
        seconds_between(
            game_finished_at,
            post_mcq_completed_at,
        )
    )

    time_post_mcq_to_post_questionnaire_sec = (
        seconds_between(
            post_mcq_completed_at,
            post_questionnaire_completed_at,
        )
    )

    # -------------------------
    # Final participant row
    # -------------------------

    row: dict[str, Any] = {
        "participant_id":
            data.get(
                "participant_id",
                np.nan,
            ),

        "meaningful_completion":
            meaningful_completion,

        "meaningful_completion_failure_reason":
            meaningful_completion_failure_reason,

        "age_years":
            demographics.get(
                "age_years",
                np.nan,
            ),

        "sex":
            demographics.get(
                "sex",
                np.nan,
            ),

        "gaming_frequency":
            demographics.get(
                "gaming_frequency",
                np.nan,
            ),

        "pre_mcq_attempted":
            pre_mcq_attempted,

        "pre_mcq_questions_attempted":
            pre_mcq_questions_attempted,

        "pre_mcq_questions_correct":
            pre_mcq_questions_correct,

        "pre_mcq_total_correct":
            pre_mcq_total_correct,

        "post_mcq_attempted":
            post_mcq_attempted,

        "post_mcq_questions_attempted":
            post_mcq_questions_attempted,

        "post_mcq_questions_correct":
            post_mcq_questions_correct,

        "post_mcq_total_correct":
            post_mcq_total_correct,

        "mcq_correct_change":
            mcq_correct_change,

        "mri_anxiety_pre":
            mri_anxiety_pre,

        "mri_anxiety_post":
            mri_anxiety_post,

        "mri_anxiety_change":
            mri_anxiety_change,

        "mri_willingness_pre":
            mri_willingness_pre,

        "mri_willingness_post":
            mri_willingness_post,

        "mri_willingness_change":
            mri_willingness_change,

        "technical_issues":
            post_questionnaire.get(
                "technical_issues",
                np.nan,
            ),

        "technical_issues_text":
            post_questionnaire.get(
                "technical_issues_text",
                np.nan,
            ),

        "most_useful_learning":
            post_questionnaire.get(
                "most_useful_learning",
                np.nan,
            ),

        "improvements":
            post_questionnaire.get(
                "improvements",
                np.nan,
            ),

        "aim_1":
            aim.get(
                "aim_1",
                np.nan,
            ),

        "aim_2":
            aim.get(
                "aim_2",
                np.nan,
            ),

        "aim_3":
            aim.get(
                "aim_3",
                np.nan,
            ),

        "aim_4":
            aim.get(
                "aim_4",
                np.nan,
            ),

        "time_demographics_to_pre_mcq_sec":
            time_demographics_to_pre_mcq_sec,

        "time_gameplay_sec":
            time_gameplay_sec,

        "time_game_finish_to_post_mcq_sec":
            time_game_finish_to_post_mcq_sec,

        "unique_npcs_interacted_with":
            unique_npcs_interacted_with,

        "npcs_interacted_with":
            npcs_interacted_with,

        "time_post_mcq_to_post_questionnaire_sec":
            time_post_mcq_to_post_questionnaire_sec,

    }

    # Add readable minutes-and-seconds columns.
    for column in TIME_COLUMNS_SEC:
        min_sec_column = column.replace(
            "_sec",
            "_min_sec",
        )

        row[min_sec_column] = (
            format_seconds_as_min_sec(
                row.get(column)
            )
        )

    return row


# =============================
# Load all participant JSON files
# =============================

def load_all_participants(
    data_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Read every .json file in the data directory.

    Non-JSON files are ignored. Files that cannot be processed are
    recorded in the processing_issues dataframe.
    """
    participant_rows: list[
        dict[str, Any]
    ] = []

    processing_issues: list[
        dict[str, str]
    ] = []

    json_files = sorted(
        data_dir.glob("*.json")
    )

    for json_file in json_files:
        try:
            with json_file.open(
                "r",
                encoding="utf-8",
            ) as file:
                participant_data = json.load(
                    file
                )

            if not isinstance(
                participant_data,
                dict,
            ):
                raise ValueError(
                    "Top-level JSON structure is not "
                    "an object/dictionary."
                )

            participant_row = (
                extract_participant_row(
                    participant_data
                )
            )

            participant_rows.append(
                participant_row
            )

        except Exception as exc:
            processing_issues.append({
                "source_file":
                    json_file.name,

                "error":
                    repr(exc),
            })

    participant_df = pd.DataFrame(
        participant_rows
    )

    issues_df = pd.DataFrame(
        processing_issues
    )

    if participant_df.empty:
        return (
            participant_df,
            issues_df,
        )

    # Build the final column order.
    final_columns: list[str] = []

    for column in REQUESTED_COLUMNS:
        final_columns.append(
            column
        )

        # Insert the two NPC fields directly after the numeric duration
        # from game completion to post-MCQ completion.
        if (
            column
            == "time_game_finish_to_post_mcq_sec"
        ):
            final_columns.append(
                "unique_npcs_interacted_with"
            )

            final_columns.append(
                "npcs_interacted_with"
            )

        # Add the readable time column after any specially inserted
        # columns associated with that seconds variable.
        if column in TIME_COLUMNS_SEC:
            min_sec_column = column.replace(
                "_sec",
                "_min_sec",
            )

            final_columns.append(
                min_sec_column
            )

    final_columns = [
        column
        for column in final_columns
        if column in participant_df.columns
    ]

    participant_df = participant_df[
        final_columns
    ]

    return (
        participant_df,
        issues_df,
    )


# =============================
# Excel formatting
# =============================

def autosize_columns(
    worksheet,
    dataframe: pd.DataFrame,
) -> None:
    """
    Adjust Excel column widths while preventing long free-text
    responses from producing excessively wide columns.
    """
    wider_text_columns = {
        "most_useful_learning",
        "improvements",
        "technical_issues_text",
        "npcs_interacted_with",
    }

    mcq_list_columns = {
        "pre_mcq_questions_attempted",
        "pre_mcq_questions_correct",
        "post_mcq_questions_attempted",
        "post_mcq_questions_correct",
    }

    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        maximum_length = len(
            str(column_name)
        )

        for value in dataframe[
            column_name
        ].head(200):
            if pd.notna(value):
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        if column_name in wider_text_columns:
            width = min(
                max(
                    maximum_length + 2,
                    25,
                ),
                60,
            )

        elif column_name in mcq_list_columns:
            width = min(
                max(
                    maximum_length + 2,
                    18,
                ),
                35,
            )

        elif (
            column_name
            == "meaningful_completion_failure_reason"
        ):
            width = min(
                max(
                    maximum_length + 2,
                    25,
                ),
                45,
            )

        else:
            width = min(
                max(
                    maximum_length + 2,
                    10,
                ),
                28,
            )

        worksheet.column_dimensions[
            column_letter
        ].width = width


def format_excel_sheet(
    writer: pd.ExcelWriter,
    dataframe: pd.DataFrame,
) -> None:
    """
    Apply formatting to the participant_summary worksheet.
    """
    worksheet = writer.sheets[
        "participant_summary"
    ]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # Highlight rows that did not satisfy meaningful-completion criteria.
    if (
        "meaningful_completion"
        in dataframe.columns
    ):
        meaningful_column_index = (
            dataframe.columns.get_loc(
                "meaningful_completion"
            )
            + 1
        )

        non_completion_fill = PatternFill(
            start_color="F8D7DA",
            end_color="F8D7DA",
            fill_type="solid",
        )

        for excel_row_index in range(
            2,
            len(dataframe) + 2,
        ):
            value = worksheet.cell(
                row=excel_row_index,
                column=meaningful_column_index,
            ).value

            if (
                value is False
                or str(value).lower() == "false"
            ):
                for excel_column_index in range(
                    1,
                    len(dataframe.columns) + 1,
                ):
                    worksheet.cell(
                        row=excel_row_index,
                        column=excel_column_index,
                    ).fill = (
                        non_completion_fill
                    )

    # Numeric display formatting.
    for column_index, column_name in enumerate(
        dataframe.columns,
        start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        if column_name.endswith("_sec"):
            for cell in worksheet[
                column_letter
            ][1:]:
                cell.number_format = "0.0"

        if column_name.endswith(
            "_change"
        ):
            for cell in worksheet[
                column_letter
            ][1:]:
                cell.number_format = "0.0"

    autosize_columns(
        worksheet,
        dataframe,
    )

    worksheet.row_dimensions[
        1
    ].height = 35

    for row_index in range(
        2,
        len(dataframe) + 2,
    ):
        worksheet.row_dimensions[
            row_index
        ].height = 45


def format_readme_sheet(
    worksheet,
) -> None:
    """
    Apply formatting to the README worksheet.
    """
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.column_dimensions[
        "A"
    ].width = 42

    worksheet.column_dimensions[
        "B"
    ].width = 100

    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


# =============================
# Write Excel workbook
# =============================

def write_excel(
    dataframe: pd.DataFrame,
    issues_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """
    Write participant summary and documentation to Excel.
    """
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    readme_df = pd.DataFrame({
        "item": [
            "meaningful_completion",
            "meaningful_completion_failure_reason",

            "pre_mcq_attempted",
            "pre_mcq_questions_attempted",
            "pre_mcq_questions_correct",

            "post_mcq_attempted",
            "post_mcq_questions_attempted",
            "post_mcq_questions_correct",

            "unique_npcs_interacted_with",
            "npcs_interacted_with",

            "time_gameplay_sec",
            "time columns ending in _sec",
            "time columns ending in _min_sec",
        ],

        "definition": [
            (
                f"True if the participant interacted with at least "
                f"{MEANINGFUL_MIN_NPCS} unique NPCs, attempted at least "
                f"{MEANINGFUL_MIN_PRE_MCQ_ATTEMPTED}/"
                f"{N_MCQ_QUESTIONS} pre-game MCQ questions, and "
                f"attempted at least "
                f"{MEANINGFUL_MIN_POST_MCQ_ATTEMPTED}/"
                f"{N_MCQ_QUESTIONS} post-game MCQ questions."
            ),

            (
                "Identifies any meaningful-completion criterion or "
                "criteria that were not satisfied."
            ),

            (
                "Number of pre-game MCQ questions answered, regardless "
                "of whether the responses were correct."
            ),

            (
                "Nominal pre-game MCQ question numbers that were "
                "attempted, listed in Q1-Q7 order."
            ),

            (
                "Nominal pre-game MCQ question numbers that were "
                "answered correctly, listed in Q1-Q7 order."
            ),

            (
                "Number of post-game MCQ questions answered, regardless "
                "of whether the responses were correct."
            ),

            (
                "Nominal post-game MCQ question numbers that were "
                "attempted, listed in Q1-Q7 order."
            ),

            (
                "Nominal post-game MCQ question numbers that were "
                "answered correctly, listed in Q1-Q7 order."
            ),

            (
                "Number of predefined NPCs with at least one recorded "
                "conversation interaction."
            ),

            (
                "Reader-facing names of the predefined NPCs with at "
                "least one recorded conversation interaction. "
                "Conversation frequencies are not shown."
            ),

            (
                "Gameplay duration in seconds, extracted directly from "
                "the game telemetry field time_played_seconds."
            ),

            (
                "Duration variables for the final analysis, expressed "
                "numerically in seconds."
            ),

            (
                "The corresponding duration displayed in a readable "
                "minutes-and-seconds format for manual inspection."
            ),
        ],
    })

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        dataframe.to_excel(
            writer,
            sheet_name="participant_summary",
            index=False,
        )

        readme_df.to_excel(
            writer,
            sheet_name="README",
            index=False,
        )

        if not issues_df.empty:
            issues_df.to_excel(
                writer,
                sheet_name="processing_issues",
                index=False,
            )

        format_excel_sheet(
            writer,
            dataframe,
        )

        format_readme_sheet(
            writer.sheets["README"]
        )


# =============================
# Main
# =============================

def main() -> None:
    if not DATA_DIR.exists():
        raise FileNotFoundError(
            f"Could not find data directory: "
            f"{DATA_DIR.resolve()}"
        )

    participant_df, issues_df = (
        load_all_participants(
            DATA_DIR
        )
    )

    if participant_df.empty:
        raise RuntimeError(
            "No valid participant JSON files were found in: "
            f"{DATA_DIR.resolve()}"
        )

    write_excel(
        participant_df,
        issues_df,
        OUTPUT_EXCEL_FILE,
    )

    print(
        "Excel summary complete."
    )

    print(
        f"Participants processed: "
        f"{len(participant_df)}"
    )

    print(
        f"Output file: "
        f"{OUTPUT_EXCEL_FILE.resolve()}"
    )

    if not issues_df.empty:
        print(
            f"Warning: {len(issues_df)} "
            "file(s) had processing issues."
        )

        print(
            "See the processing_issues worksheet "
            "in the Excel file."
        )


if __name__ == "__main__":
    main()