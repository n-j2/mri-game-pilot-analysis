
from __future__ import annotations

"""
Final analysis script for the MRI educational game pilot study.

Expected input
--------------
analysis_output/participant_summary.xlsx
Worksheet: participant_summary

Main outputs
------------
analysis_output/final_analysis/
    final_analysis_results.xlsx
    final_analysis_report.txt
    qualitative_coding_template.xlsx

The script implements:
- sample-flow and missing-data summaries
- demographic descriptives
- protocol Go/Amend/Stop feasibility criteria
- primary paired analysis of MCQ knowledge score
- secondary paired analyses of MRI anxiety and willingness
- exact conditional-permutation Wilcoxon signed-rank tests
- exact sign-test sensitivity analyses
- Hodges-Lehmann paired shifts
- matched-pairs rank-biserial correlations
- reproducible bootstrap 95% confidence intervals
- Holm adjustment across the two secondary outcomes
- per-item MCQ accuracy and exploratory exact McNemar tests
- AIM scoring, descriptive statistics, and exploratory Cronbach alpha
- technical-issue and telemetry summaries
- a de-linked workbook for manual qualitative coding

Required packages
-----------------
pandas, numpy, scipy, openpyxl

Install, if needed:
    pip install pandas numpy scipy openpyxl
"""

import math
import platform
import re
import sys
from pathlib import Path
from typing import Any, Callable

import numpy as np
import pandas as pd
import scipy
from scipy.stats import binomtest, norm, rankdata, skew, t
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# User settings
# =============================================================================

INPUT_FILE = Path("analysis_output") / "participant_summary.xlsx"
INPUT_SHEET = "participant_summary"

OUTPUT_DIR = Path("analysis_output") / "final_analysis"
RESULTS_WORKBOOK = OUTPUT_DIR / "final_analysis_results.xlsx"
TEXT_REPORT = OUTPUT_DIR / "final_analysis_report.txt"
QUALITATIVE_WORKBOOK = OUTPUT_DIR / "qualitative_coding_template.xlsx"

# Protocol-defined thresholds
MCQ_MIN_ATTEMPTED = 5
N_MCQ_ITEMS = 7
SHORT_GAMEPLAY_SECONDS = 4 * 60
LONG_GAMEPLAY_SECONDS = 40 * 60

# AIM participant mean is calculated only when this many items are answered.
# Requiring all four items is conservative and easy to explain.
AIM_MIN_ITEMS_FOR_MEAN = 4

# Reproducible bootstrap settings
BOOTSTRAP_REPLICATES = 10_000
BOOTSTRAP_SEED = 20260804
CONFIDENCE_LEVEL = 0.95

# Optional sensitivity analysis restricted to meaningful completers
RUN_MEANINGFUL_COMPLETER_SENSITIVITY = True


# =============================================================================
# Expected variables and response mappings
# =============================================================================

REQUIRED_COLUMNS = [
    "participant_id",
    "meaningful_completion",
    "age_years",
    "sex",
    "gaming_frequency",
    "pre_mcq_attempted",
    "pre_mcq_total_correct",
    "pre_mcq_questions_correct",
    "post_mcq_attempted",
    "post_mcq_total_correct",
    "post_mcq_questions_correct",
    "mri_anxiety_pre",
    "mri_anxiety_post",
    "mri_willingness_pre",
    "mri_willingness_post",
    "time_gameplay_sec",
    "unique_npcs_interacted_with",
    "technical_issues",
    "aim_1",
    "aim_2",
    "aim_3",
    "aim_4",
    "most_useful_learning",
    "improvements",
]

AIM_COLUMNS = ["aim_1", "aim_2", "aim_3", "aim_4"]

AIM_SCORE_MAP = {
    "completely_disagree": 1,
    "strongly_disagree": 1,
    "disagree": 2,
    "neither_agree_nor_disagree": 3,
    "neither_disagree_nor_agree": 3,
    "neutral": 3,
    "neither": 3,
    "agree": 4,
    "completely_agree": 5,
    "strongly_agree": 5,
}

AIM_RESPONSE_LABELS = {
    1: "Completely disagree",
    2: "Disagree",
    3: "Neither agree nor disagree",
    4: "Agree",
    5: "Completely agree",
}

TECHNICAL_ISSUE_MAP = {
    "no": "No",
    "minor_problems": "Minor problems",
    "yes_minor": "Minor problems",
    "minor": "Minor problems",
    "major_problems": "Major problems",
    "yes_major": "Major problems",
    "major": "Major problems",
}

OUTCOMES = [
    {
        "outcome": "MCQ knowledge score",
        "role": "Primary",
        "pre": "pre_mcq_total_correct",
        "post": "post_mcq_total_correct",
        "favourable": "increase",
        "mask": "mcq_usable",
        "scale": "0-7",
    },
    {
        "outcome": "MRI anxiety",
        "role": "Secondary",
        "pre": "mri_anxiety_pre",
        "post": "mri_anxiety_post",
        "favourable": "decrease",
        "mask": "anxiety_pair",
        "scale": "0-10",
    },
    {
        "outcome": "MRI willingness",
        "role": "Secondary",
        "pre": "mri_willingness_pre",
        "post": "mri_willingness_post",
        "favourable": "increase",
        "mask": "willingness_pair",
        "scale": "0-10",
    },
]


# =============================================================================
# General helpers
# =============================================================================

def load_data() -> pd.DataFrame:
    """Load the participant summary workbook."""
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Could not find: {INPUT_FILE.resolve()}\n"
            "Run publication_make_excel_summary.py first."
        )

    df = pd.read_excel(INPUT_FILE, sheet_name=INPUT_SHEET)

    if df.empty:
        raise RuntimeError("The participant_summary worksheet contains no rows.")

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        raise RuntimeError(
            "The following required columns are missing:\n- "
            + "\n- ".join(missing)
        )

    if df["participant_id"].duplicated().any():
        duplicates = df.loc[
            df["participant_id"].duplicated(keep=False),
            "participant_id",
        ].astype(str).tolist()
        raise RuntimeError(
            "Duplicate participant IDs were found: " + ", ".join(duplicates)
        )

    return df.copy()


def normalise_text(value: Any) -> str:
    """Convert categorical text to lowercase underscore format."""
    if pd.isna(value):
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
    )


def parse_boolean(value: Any) -> bool | None:
    """Interpret common Boolean representations."""
    if isinstance(value, (bool, np.bool_)):
        return bool(value)
    if pd.isna(value):
        return None

    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def numeric(series: pd.Series) -> pd.Series:
    """Convert a Series to numeric values; invalid entries become missing."""
    return pd.to_numeric(series, errors="coerce")


def quantile(values: np.ndarray, probability: float) -> float:
    """Return a quantile, or NaN for an empty array."""
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        return np.nan
    return float(np.quantile(values, probability))


def continuous_summary(values: pd.Series, variable: str) -> dict[str, Any]:
    """Standard descriptive statistics for a continuous variable."""
    array = numeric(values).dropna().to_numpy(dtype=float)

    if array.size == 0:
        return {
            "variable": variable,
            "n": 0,
            "mean": np.nan,
            "sd": np.nan,
            "median": np.nan,
            "q1": np.nan,
            "q3": np.nan,
            "minimum": np.nan,
            "maximum": np.nan,
        }

    return {
        "variable": variable,
        "n": int(array.size),
        "mean": float(np.mean(array)),
        "sd": float(np.std(array, ddof=1)) if array.size > 1 else np.nan,
        "median": float(np.median(array)),
        "q1": quantile(array, 0.25),
        "q3": quantile(array, 0.75),
        "minimum": float(np.min(array)),
        "maximum": float(np.max(array)),
    }


def wilson_ci(successes: int, total: int) -> tuple[float, float]:
    """Wilson 95% confidence interval for a binomial proportion."""
    if total <= 0:
        return np.nan, np.nan

    alpha = 1 - CONFIDENCE_LEVEL
    z = float(norm.ppf(1 - alpha / 2))
    p = successes / total

    denominator = 1 + z**2 / total
    centre = (p + z**2 / (2 * total)) / denominator
    half_width = (
        z
        * math.sqrt(
            p * (1 - p) / total
            + z**2 / (4 * total**2)
        )
        / denominator
    )

    return max(0.0, centre - half_width), min(1.0, centre + half_width)


def mean_ci(values: np.ndarray) -> tuple[float, float]:
    """Two-sided t confidence interval for a mean."""
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]

    if values.size < 2:
        return np.nan, np.nan

    mean_value = float(np.mean(values))
    standard_error = float(np.std(values, ddof=1) / math.sqrt(values.size))
    alpha = 1 - CONFIDENCE_LEVEL
    critical = float(t.ppf(1 - alpha / 2, df=values.size - 1))

    return (
        mean_value - critical * standard_error,
        mean_value + critical * standard_error,
    )


def holm_adjust(p_values: list[float]) -> list[float]:
    """Holm step-down multiple-testing adjustment."""
    p_array = np.asarray(p_values, dtype=float)
    adjusted = np.full_like(p_array, np.nan, dtype=float)
    valid = np.flatnonzero(np.isfinite(p_array))

    if valid.size == 0:
        return adjusted.tolist()

    ordered = valid[np.argsort(p_array[valid])]
    number_of_tests = len(ordered)
    running_maximum = 0.0

    for position, original_index in enumerate(ordered):
        multiplier = number_of_tests - position
        candidate = min(1.0, p_array[original_index] * multiplier)
        running_maximum = max(running_maximum, candidate)
        adjusted[original_index] = running_maximum

    return adjusted.tolist()


def format_p(value: Any) -> str:
    """Publication-style p-value formatting."""
    if value is None or pd.isna(value):
        return "NA"
    value = float(value)
    if value < 0.001:
        return "<.001"
    return f"{value:.3f}".lstrip("0")


# =============================================================================
# Input validation and analysis populations
# =============================================================================

def validate_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Check expected value ranges and stored change-score consistency.

    Out-of-range values stop the analysis. Change-score discrepancies are
    recorded as warnings because the script recalculates all changes from
    the pre and post values.
    """
    issues: list[dict[str, Any]] = []

    ranges = {
        "pre_mcq_attempted": (0, 7),
        "post_mcq_attempted": (0, 7),
        "pre_mcq_total_correct": (0, 7),
        "post_mcq_total_correct": (0, 7),
        "mri_anxiety_pre": (0, 10),
        "mri_anxiety_post": (0, 10),
        "mri_willingness_pre": (0, 10),
        "mri_willingness_post": (0, 10),
        "unique_npcs_interacted_with": (0, 12),
    }

    for column, (minimum, maximum) in ranges.items():
        values = numeric(df[column])
        invalid = values.notna() & (
            (values < minimum) | (values > maximum)
        )

        for index in df.index[invalid]:
            issues.append(
                {
                    "severity": "error",
                    "participant_id": df.loc[index, "participant_id"],
                    "column": column,
                    "issue": f"Value outside expected range {minimum}-{maximum}",
                    "observed_value": df.loc[index, column],
                }
            )

    for prefix in ["pre", "post"]:
        attempted = numeric(df[f"{prefix}_mcq_attempted"])
        correct = numeric(df[f"{prefix}_mcq_total_correct"])
        invalid = attempted.notna() & correct.notna() & (correct > attempted)

        for index in df.index[invalid]:
            issues.append(
                {
                    "severity": "error",
                    "participant_id": df.loc[index, "participant_id"],
                    "column": f"{prefix}_mcq_total_correct",
                    "issue": "Correct total exceeded attempted total",
                    "observed_value": correct.loc[index],
                }
            )

    nonnegative_columns = [
        "time_gameplay_sec",
        "time_demographics_to_pre_mcq_sec",
        "time_game_finish_to_post_mcq_sec",
        "time_post_mcq_to_post_questionnaire_sec",
        "time_total_session_sec",
    ]

    for column in nonnegative_columns:
        if column not in df.columns:
            continue
        values = numeric(df[column])
        invalid = values.notna() & (values < 0)

        for index in df.index[invalid]:
            issues.append(
                {
                    "severity": "error",
                    "participant_id": df.loc[index, "participant_id"],
                    "column": column,
                    "issue": "Negative duration",
                    "observed_value": df.loc[index, column],
                }
            )

    change_checks = [
        ("mcq_correct_change", "pre_mcq_total_correct", "post_mcq_total_correct"),
        ("mri_anxiety_change", "mri_anxiety_pre", "mri_anxiety_post"),
        (
            "mri_willingness_change",
            "mri_willingness_pre",
            "mri_willingness_post",
        ),
    ]

    for change_column, pre_column, post_column in change_checks:
        if change_column not in df.columns:
            continue

        stored = numeric(df[change_column])
        recalculated = numeric(df[post_column]) - numeric(df[pre_column])
        mismatch = (
            stored.notna()
            & recalculated.notna()
            & ~np.isclose(stored, recalculated)
        )

        for index in df.index[mismatch]:
            issues.append(
                {
                    "severity": "warning",
                    "participant_id": df.loc[index, "participant_id"],
                    "column": change_column,
                    "issue": (
                        "Stored change did not equal post minus pre; "
                        "recalculated value used"
                    ),
                    "observed_value": stored.loc[index],
                }
            )

    issues_df = pd.DataFrame(issues)

    if (
        not issues_df.empty
        and (issues_df["severity"] == "error").any()
    ):
        raise RuntimeError(
            "Input validation failed:\n\n"
            + issues_df.to_string(index=False)
        )

    return issues_df


def build_masks(df: pd.DataFrame) -> dict[str, pd.Series]:
    """Build protocol-consistent analysis-population masks."""
    pre_attempted = numeric(df["pre_mcq_attempted"])
    post_attempted = numeric(df["post_mcq_attempted"])

    mcq_usable = (
        pre_attempted.ge(MCQ_MIN_ATTEMPTED)
        & post_attempted.ge(MCQ_MIN_ATTEMPTED)
        & numeric(df["pre_mcq_total_correct"]).notna()
        & numeric(df["post_mcq_total_correct"]).notna()
    )

    anxiety_pair = (
        numeric(df["mri_anxiety_pre"]).notna()
        & numeric(df["mri_anxiety_post"]).notna()
    )

    willingness_pair = (
        numeric(df["mri_willingness_pre"]).notna()
        & numeric(df["mri_willingness_post"]).notna()
    )

    telemetry_usable = (
        numeric(df["time_gameplay_sec"]).notna()
        & numeric(df["unique_npcs_interacted_with"]).notna()
    )

    meaningful_values = df["meaningful_completion"].apply(parse_boolean)

    if meaningful_values.isna().any():
        affected = df.loc[
            meaningful_values.isna(),
            "participant_id",
        ].astype(str).tolist()
        raise RuntimeError(
            "Could not interpret meaningful_completion for: "
            + ", ".join(affected)
        )

    meaningful = meaningful_values.astype(bool)

    fully_usable = (
        mcq_usable
        & anxiety_pair
        & willingness_pair
        & telemetry_usable
    )

    return {
        "all": pd.Series(True, index=df.index),
        "mcq_usable": mcq_usable,
        "anxiety_pair": anxiety_pair,
        "willingness_pair": willingness_pair,
        "telemetry_usable": telemetry_usable,
        "meaningful": meaningful,
        "fully_usable": fully_usable,
    }


# =============================================================================
# Exact paired nonparametric analyses
# =============================================================================

def signed_rank_components(
    differences: np.ndarray,
) -> dict[str, float]:
    """
    Wilcoxon signed-rank components.

    Zero differences are removed before ranking. Average ranks are used for
    tied absolute differences.
    """
    differences = np.asarray(differences, dtype=float)
    differences = differences[np.isfinite(differences)]
    nonzero = differences[~np.isclose(differences, 0)]

    if nonzero.size == 0:
        return {
            "nonzero_n": 0,
            "w_plus": 0.0,
            "w_minus": 0.0,
            "w": 0.0,
            "rank_biserial": 0.0,
        }

    ranks = rankdata(np.abs(nonzero), method="average")
    w_plus = float(np.sum(ranks[nonzero > 0]))
    w_minus = float(np.sum(ranks[nonzero < 0]))
    total = w_plus + w_minus

    return {
        "nonzero_n": int(nonzero.size),
        "w_plus": w_plus,
        "w_minus": w_minus,
        "w": min(w_plus, w_minus),
        "rank_biserial": (
            (w_plus - w_minus) / total
            if total > 0
            else 0.0
        ),
    }


def exact_signed_rank_p(differences: np.ndarray) -> float:
    """
    Exact two-sided conditional-permutation Wilcoxon signed-rank p value.

    Dynamic programming enumerates the null distribution of all possible
    sign allocations. This handles tied absolute differences and avoids a
    large-sample normal approximation.
    """
    differences = np.asarray(differences, dtype=float)
    differences = differences[np.isfinite(differences)]
    nonzero = differences[~np.isclose(differences, 0)]

    if nonzero.size == 0:
        return 1.0

    ranks = rankdata(np.abs(nonzero), method="average")
    scaled_ranks = np.rint(ranks * 2).astype(int)

    observed_positive_sum = int(
        np.sum(scaled_ranks[nonzero > 0])
    )
    total_rank_sum = int(np.sum(scaled_ranks))

    counts = [0] * (total_rank_sum + 1)
    counts[0] = 1
    reachable = 0

    for rank_value in scaled_ranks:
        for current_sum in range(reachable, -1, -1):
            if counts[current_sum]:
                counts[current_sum + rank_value] += counts[current_sum]
        reachable += rank_value

    observed_extremeness = abs(
        2 * observed_positive_sum - total_rank_sum
    )

    extreme_assignments = sum(
        assignment_count
        for rank_sum, assignment_count in enumerate(counts)
        if abs(2 * rank_sum - total_rank_sum)
        >= observed_extremeness
    )

    return min(
        1.0,
        extreme_assignments / (2 ** int(nonzero.size)),
    )


def hodges_lehmann_shift(differences: np.ndarray) -> float:
    """
    Paired Hodges-Lehmann location-shift estimate.

    This is the median of all Walsh averages:
        (d_i + d_j) / 2, for i <= j.
    """
    differences = np.asarray(differences, dtype=float)
    differences = differences[np.isfinite(differences)]

    if differences.size == 0:
        return np.nan

    walsh_matrix = (
        differences[:, None] + differences[None, :]
    ) / 2
    walsh_values = walsh_matrix[
        np.triu_indices(differences.size)
    ]

    return float(np.median(walsh_values))


def rank_biserial(differences: np.ndarray) -> float:
    """Matched-pairs rank-biserial correlation."""
    return float(
        signed_rank_components(differences)["rank_biserial"]
    )


def bootstrap_ci(
    differences: np.ndarray,
    statistic: Callable[[np.ndarray], float],
    seed: int,
) -> tuple[float, float]:
    """
    Percentile bootstrap confidence interval using paired-case resampling.
    """
    differences = np.asarray(differences, dtype=float)
    differences = differences[np.isfinite(differences)]

    if differences.size < 2:
        return np.nan, np.nan

    rng = np.random.default_rng(seed)
    n = differences.size
    bootstrap_values = np.empty(
        BOOTSTRAP_REPLICATES,
        dtype=float,
    )

    for replicate in range(BOOTSTRAP_REPLICATES):
        indices = rng.integers(0, n, size=n)
        bootstrap_values[replicate] = statistic(
            differences[indices]
        )

    bootstrap_values = bootstrap_values[
        np.isfinite(bootstrap_values)
    ]

    alpha = 1 - CONFIDENCE_LEVEL
    return (
        float(np.quantile(bootstrap_values, alpha / 2)),
        float(np.quantile(bootstrap_values, 1 - alpha / 2)),
    )


def analyse_outcome(
    df: pd.DataFrame,
    spec: dict[str, Any],
    eligibility_mask: pd.Series,
    population_label: str,
    seed_offset: int,
) -> dict[str, Any]:
    """Analyse one paired outcome."""
    pre = numeric(df[spec["pre"]])
    post = numeric(df[spec["post"]])
    valid = eligibility_mask & pre.notna() & post.notna()

    pre_values = pre.loc[valid].to_numpy(dtype=float)
    post_values = post.loc[valid].to_numpy(dtype=float)
    differences = post_values - pre_values

    if differences.size == 0:
        raise RuntimeError(
            f"No paired data were available for {spec['outcome']}."
        )

    components = signed_rank_components(differences)

    positive = int(np.sum(differences > 0))
    negative = int(np.sum(differences < 0))
    zero = int(np.sum(np.isclose(differences, 0)))
    non_tied = positive + negative

    sign_p = (
        float(
            binomtest(
                positive,
                non_tied,
                p=0.5,
                alternative="two-sided",
            ).pvalue
        )
        if non_tied > 0
        else 1.0
    )

    median_ci = bootstrap_ci(
        differences,
        statistic=lambda x: float(np.median(x)),
        seed=BOOTSTRAP_SEED + seed_offset,
    )

    hl_ci = bootstrap_ci(
        differences,
        statistic=hodges_lehmann_shift,
        seed=BOOTSTRAP_SEED + 100 + seed_offset,
    )

    rb_ci = bootstrap_ci(
        differences,
        statistic=rank_biserial,
        seed=BOOTSTRAP_SEED + 200 + seed_offset,
    )

    if spec["favourable"] == "increase":
        favourable = positive
        unfavourable = negative
    else:
        favourable = negative
        unfavourable = positive

    return {
        "analysis_population": population_label,
        "outcome": spec["outcome"],
        "analysis_role": spec["role"],
        "scale": spec["scale"],
        "favourable_direction": spec["favourable"],
        "paired_n": int(differences.size),

        "pre_mean": float(np.mean(pre_values)),
        "pre_sd": (
            float(np.std(pre_values, ddof=1))
            if pre_values.size > 1
            else np.nan
        ),
        "pre_median": float(np.median(pre_values)),
        "pre_q1": quantile(pre_values, 0.25),
        "pre_q3": quantile(pre_values, 0.75),
        "pre_minimum": float(np.min(pre_values)),
        "pre_maximum": float(np.max(pre_values)),

        "post_mean": float(np.mean(post_values)),
        "post_sd": (
            float(np.std(post_values, ddof=1))
            if post_values.size > 1
            else np.nan
        ),
        "post_median": float(np.median(post_values)),
        "post_q1": quantile(post_values, 0.25),
        "post_q3": quantile(post_values, 0.75),
        "post_minimum": float(np.min(post_values)),
        "post_maximum": float(np.max(post_values)),

        "change_mean": float(np.mean(differences)),
        "change_sd": (
            float(np.std(differences, ddof=1))
            if differences.size > 1
            else np.nan
        ),
        "change_median": float(np.median(differences)),
        "change_q1": quantile(differences, 0.25),
        "change_q3": quantile(differences, 0.75),
        "change_minimum": float(np.min(differences)),
        "change_maximum": float(np.max(differences)),
        "change_skewness": (
            float(skew(differences, bias=False))
            if differences.size >= 3
            else np.nan
        ),

        "n_increased": positive,
        "n_unchanged": zero,
        "n_decreased": negative,
        "n_favourable_change": favourable,
        "n_unfavourable_change": unfavourable,

        "wilcoxon_nonzero_n": components["nonzero_n"],
        "wilcoxon_w_plus": components["w_plus"],
        "wilcoxon_w_minus": components["w_minus"],
        "wilcoxon_w_two_sided": components["w"],
        "wilcoxon_exact_permutation_p": exact_signed_rank_p(
            differences
        ),

        "sign_test_non_tied_n": non_tied,
        "sign_test_exact_p": sign_p,

        "median_change_ci_lower": median_ci[0],
        "median_change_ci_upper": median_ci[1],

        "hodges_lehmann_shift": hodges_lehmann_shift(
            differences
        ),
        "hodges_lehmann_ci_lower": hl_ci[0],
        "hodges_lehmann_ci_upper": hl_ci[1],

        "rank_biserial": components["rank_biserial"],
        "rank_biserial_ci_lower": rb_ci[0],
        "rank_biserial_ci_upper": rb_ci[1],

        "holm_adjusted_p_secondary": np.nan,
    }


def run_outcome_analyses(
    df: pd.DataFrame,
    masks: dict[str, pd.Series],
    meaningful_only: bool,
) -> pd.DataFrame:
    """Run all three paired outcome analyses."""
    rows = []

    for index, spec in enumerate(OUTCOMES):
        mask = masks[spec["mask"]].copy()

        if meaningful_only:
            mask &= masks["meaningful"]
            population_label = (
                "Meaningful completers only (sensitivity)"
            )
        else:
            population_label = (
                "Protocol-consistent available-case"
            )

        rows.append(
            analyse_outcome(
                df=df,
                spec=spec,
                eligibility_mask=mask,
                population_label=population_label,
                seed_offset=index,
            )
        )

    results = pd.DataFrame(rows)

    secondary_mask = (
        results["analysis_role"] == "Secondary"
    )

    adjusted = holm_adjust(
        results.loc[
            secondary_mask,
            "wilcoxon_exact_permutation_p",
        ].tolist()
    )

    results.loc[
        secondary_mask,
        "holm_adjusted_p_secondary",
    ] = adjusted

    return results


# =============================================================================
# Descriptive and feasibility summaries
# =============================================================================

def sample_flow(
    df: pd.DataFrame,
    masks: dict[str, pd.Series],
) -> pd.DataFrame:
    """Sample sizes used in each analysis."""
    total = len(df)

    rows = [
        (
            "Participants in participant summary",
            total,
        ),
        (
            "Meaningful completers",
            int(masks["meaningful"].sum()),
        ),
        (
            "Usable paired MCQ sample",
            int(masks["mcq_usable"].sum()),
        ),
        (
            "Usable paired anxiety sample",
            int(masks["anxiety_pair"].sum()),
        ),
        (
            "Usable paired willingness sample",
            int(masks["willingness_pair"].sum()),
        ),
        (
            "Usable gameplay telemetry",
            int(masks["telemetry_usable"].sum()),
        ),
        (
            "Fully usable linked questionnaire + telemetry dataset",
            int(masks["fully_usable"].sum()),
        ),
    ]

    output = pd.DataFrame(
        rows,
        columns=["sample_definition", "n"],
    )

    output["percentage_of_total"] = (
        output["n"] / total * 100
    )

    return output


def demographics(df: pd.DataFrame) -> pd.DataFrame:
    """Age, sex, and gaming-frequency descriptives."""
    rows: list[dict[str, Any]] = []

    age = continuous_summary(
        df["age_years"],
        "Age, years",
    )

    for statistic in [
        "n",
        "mean",
        "sd",
        "median",
        "q1",
        "q3",
        "minimum",
        "maximum",
    ]:
        rows.append(
            {
                "variable": "Age, years",
                "level_or_statistic": statistic,
                "n": (
                    age["n"]
                    if statistic == "n"
                    else np.nan
                ),
                "value": age[statistic],
                "percentage": np.nan,
            }
        )

    for column, label in [
        ("sex", "Sex"),
        ("gaming_frequency", "Gaming frequency"),
    ]:
        counts = (
            df[column]
            .fillna("Missing")
            .astype(str)
            .value_counts(dropna=False)
        )

        for level, count in counts.items():
            rows.append(
                {
                    "variable": label,
                    "level_or_statistic": level,
                    "n": int(count),
                    "value": np.nan,
                    "percentage": (
                        count / len(df) * 100
                    ),
                }
            )

    return pd.DataFrame(rows)


def missingness(df: pd.DataFrame) -> pd.DataFrame:
    """Missingness for variables used in the final analyses."""
    columns = [
        "pre_mcq_total_correct",
        "post_mcq_total_correct",
        "mri_anxiety_pre",
        "mri_anxiety_post",
        "mri_willingness_pre",
        "mri_willingness_post",
        "time_gameplay_sec",
        "unique_npcs_interacted_with",
        "technical_issues",
        *AIM_COLUMNS,
        "most_useful_learning",
        "improvements",
    ]

    rows = []
    total = len(df)

    for column in columns:
        missing_n = int(df[column].isna().sum())
        rows.append(
            {
                "variable": column,
                "total_n": total,
                "nonmissing_n": total - missing_n,
                "missing_n": missing_n,
                "missing_percentage": (
                    missing_n / total * 100
                ),
            }
        )

    return pd.DataFrame(rows)


def score_aim(
    df: pd.DataFrame,
) -> tuple[
    pd.DataFrame,
    pd.Series,
    pd.DataFrame,
]:
    """
    Convert AIM responses to 1-5 scores.

    Returns:
    - item-score dataframe
    - participant mean score
    - unrecognised-response issues
    """
    scored = pd.DataFrame(index=df.index)
    issues = []

    for column in AIM_COLUMNS:
        normalised = df[column].apply(normalise_text)
        scores = normalised.map(AIM_SCORE_MAP)

        unrecognised = (
            normalised.ne("")
            & scores.isna()
        )

        for index in df.index[unrecognised]:
            issues.append(
                {
                    "participant_id": df.loc[
                        index,
                        "participant_id",
                    ],
                    "column": column,
                    "issue": "Unrecognised AIM response",
                    "observed_value": df.loc[index, column],
                }
            )

        scored[column] = scores

    completed_items = scored.notna().sum(axis=1)

    participant_mean = scored.mean(
        axis=1,
        skipna=True,
    ).where(
        completed_items >= AIM_MIN_ITEMS_FOR_MEAN
    )

    return (
        scored,
        participant_mean,
        pd.DataFrame(issues),
    )


def cronbach_alpha(scores: pd.DataFrame) -> float:
    """Cronbach alpha using complete AIM rows."""
    complete = scores.dropna(axis=0, how="any")
    item_count = complete.shape[1]

    if complete.shape[0] < 2 or item_count < 2:
        return np.nan

    item_variance_sum = complete.var(
        axis=0,
        ddof=1,
    ).sum()

    total_variance = complete.sum(
        axis=1
    ).var(ddof=1)

    if total_variance <= 0:
        return np.nan

    return float(
        item_count
        / (item_count - 1)
        * (
            1
            - item_variance_sum / total_variance
        )
    )


def aim_outputs(
    df: pd.DataFrame,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
]:
    """Create AIM summary, item distributions, score distribution, and issues."""
    scored, participant_mean, issues = score_aim(df)
    valid_means = participant_mean.dropna().to_numpy(
        dtype=float
    )
    ci = mean_ci(valid_means)

    summary = pd.DataFrame(
        [
            {
                "scoring_rule": (
                    f"Participant mean requires at least "
                    f"{AIM_MIN_ITEMS_FOR_MEAN}/4 completed items"
                ),
                "participants_total": len(df),
                "participants_with_score": int(
                    np.isfinite(valid_means).sum()
                ),
                "mean": (
                    float(np.mean(valid_means))
                    if valid_means.size
                    else np.nan
                ),
                "sd": (
                    float(np.std(valid_means, ddof=1))
                    if valid_means.size > 1
                    else np.nan
                ),
                "mean_ci_lower": ci[0],
                "mean_ci_upper": ci[1],
                "median": (
                    float(np.median(valid_means))
                    if valid_means.size
                    else np.nan
                ),
                "q1": quantile(valid_means, 0.25),
                "q3": quantile(valid_means, 0.75),
                "minimum": (
                    float(np.min(valid_means))
                    if valid_means.size
                    else np.nan
                ),
                "maximum": (
                    float(np.max(valid_means))
                    if valid_means.size
                    else np.nan
                ),
                "cronbach_alpha_complete_cases": (
                    cronbach_alpha(scored)
                ),
                "complete_case_n_for_alpha": int(
                    scored.dropna().shape[0]
                ),
            }
        ]
    )

    distribution_rows = []

    for column in AIM_COLUMNS:
        item_nonmissing = int(scored[column].notna().sum())

        for score_value in [1, 2, 3, 4, 5]:
            count = int(
                (scored[column] == score_value).sum()
            )

            distribution_rows.append(
                {
                    "aim_item": column,
                    "response_score": score_value,
                    "response_label": (
                        AIM_RESPONSE_LABELS[score_value]
                    ),
                    "count": count,
                    "percentage_of_total_sample": (
                        count / len(df) * 100
                    ),
                    "percentage_of_item_respondents": (
                        count / item_nonmissing * 100
                        if item_nonmissing
                        else np.nan
                    ),
                }
            )

        missing_n = int(scored[column].isna().sum())

        distribution_rows.append(
            {
                "aim_item": column,
                "response_score": np.nan,
                "response_label": "Not answered",
                "count": missing_n,
                "percentage_of_total_sample": (
                    missing_n / len(df) * 100
                ),
                "percentage_of_item_respondents": np.nan,
            }
        )

    distributions = pd.DataFrame(
        distribution_rows
    )

    score_distribution = (
        participant_mean
        .dropna()
        .round(3)
        .value_counts()
        .sort_index()
        .rename_axis("participant_aim_mean")
        .reset_index(name="count")
    )

    if not score_distribution.empty:
        score_distribution["percentage"] = (
            score_distribution["count"]
            / score_distribution["count"].sum()
            * 100
        )

    return (
        summary,
        distributions,
        score_distribution,
        issues,
    )


def technical_issues(df: pd.DataFrame) -> pd.DataFrame:
    """Technical-issue categories."""
    normalised = df["technical_issues"].apply(
        normalise_text
    )

    mapped = normalised.map(
        TECHNICAL_ISSUE_MAP
    )

    mapped = mapped.where(
        normalised.ne(""),
        "Missing",
    ).fillna("Unrecognised")

    valid_respondents = int(
        (mapped != "Missing").sum()
    )

    rows = []

    for label in [
        "No",
        "Minor problems",
        "Major problems",
        "Missing",
        "Unrecognised",
    ]:
        count = int((mapped == label).sum())

        rows.append(
            {
                "category": label,
                "count": count,
                "percentage_of_total_sample": (
                    count / len(df) * 100
                ),
                "percentage_of_respondents": (
                    count / valid_respondents * 100
                    if (
                        valid_respondents > 0
                        and label != "Missing"
                    )
                    else np.nan
                ),
            }
        )

    return pd.DataFrame(rows)


def telemetry(df: pd.DataFrame) -> pd.DataFrame:
    """Gameplay and questionnaire-stage timing descriptives."""
    rows = [
        continuous_summary(
            df["time_gameplay_sec"],
            "Gameplay duration, seconds",
        ),
        continuous_summary(
            df["unique_npcs_interacted_with"],
            "Unique NPCs interacted with",
        ),
    ]

    optional_columns = [
        (
            "time_demographics_to_pre_mcq_sec",
            "Pre-game MCQ duration, seconds",
        ),
        (
            "time_game_finish_to_post_mcq_sec",
            "Post-game MCQ duration, seconds",
        ),
        (
            "time_post_mcq_to_post_questionnaire_sec",
            "Post-game AIM/feedback duration, seconds",
        ),
        (
            "time_total_session_sec",
            "Total session duration, seconds",
        ),
    ]

    for column, label in optional_columns:
        if column in df.columns:
            rows.append(
                continuous_summary(df[column], label)
            )

    return pd.DataFrame(rows)


def classify_progression(
    criterion: str,
    observed_value: float,
) -> str:
    """Apply the protocol's Go/Amend/Stop thresholds."""
    if pd.isna(observed_value):
        return "Not assessable"

    if criterion == "Recruitment":
        if observed_value >= 20:
            return "Go"
        if observed_value >= 15:
            return "Amend"
        return "Stop"

    if criterion == "Meaningful study completion":
        if observed_value >= 80:
            return "Go"
        if observed_value >= 65:
            return "Amend"
        return "Stop"

    if criterion == "Usable linked datasets":
        if observed_value >= 90:
            return "Go"
        if observed_value >= 75:
            return "Amend"
        return "Stop"

    if criterion == "AIM mean":
        if observed_value >= 4.0:
            return "Go"
        if observed_value >= 3.0:
            return "Amend"
        return "Stop"

    if criterion == "Major technical issues":
        if observed_value < 10:
            return "Go"
        if observed_value <= 25:
            return "Amend"
        return "Stop"

    return "Not assessable"


def feasibility(
    df: pd.DataFrame,
    masks: dict[str, pd.Series],
    aim_summary: pd.DataFrame,
    technical_summary: pd.DataFrame,
) -> pd.DataFrame:
    """Protocol progression-criterion table."""
    total = len(df)
    meaningful_n = int(masks["meaningful"].sum())
    usable_n = int(masks["fully_usable"].sum())

    major_n = int(
        technical_summary.loc[
            technical_summary["category"]
            == "Major problems",
            "count",
        ].iloc[0]
    )

    aim_mean = float(
        aim_summary["mean"].iloc[0]
    )

    rows = [
        {
            "criterion": "Recruitment",
            "numerator": total,
            "denominator": np.nan,
            "observed_value": float(total),
            "unit": "participants",
            "ci_lower": np.nan,
            "ci_upper": np.nan,
            "protocol_thresholds": (
                "Go >=20; Amend 15-19; Stop <15"
            ),
            "classification": classify_progression(
                "Recruitment",
                float(total),
            ),
        }
    ]

    for criterion, numerator, threshold_text in [
        (
            "Meaningful study completion",
            meaningful_n,
            "Go >=80%; Amend 65-79%; Stop <65%",
        ),
        (
            "Usable linked datasets",
            usable_n,
            "Go >=90%; Amend 75-89%; Stop <75%",
        ),
    ]:
        percentage = numerator / total * 100
        ci = wilson_ci(numerator, total)

        rows.append(
            {
                "criterion": criterion,
                "numerator": numerator,
                "denominator": total,
                "observed_value": percentage,
                "unit": "percent",
                "ci_lower": ci[0] * 100,
                "ci_upper": ci[1] * 100,
                "protocol_thresholds": threshold_text,
                "classification": classify_progression(
                    criterion,
                    percentage,
                ),
            }
        )

    rows.append(
        {
            "criterion": "AIM mean",
            "numerator": int(
                aim_summary[
                    "participants_with_score"
                ].iloc[0]
            ),
            "denominator": total,
            "observed_value": aim_mean,
            "unit": "mean score (1-5)",
            "ci_lower": float(
                aim_summary["mean_ci_lower"].iloc[0]
            ),
            "ci_upper": float(
                aim_summary["mean_ci_upper"].iloc[0]
            ),
            "protocol_thresholds": (
                "Go >=4.0; Amend 3.0-3.9; Stop <3.0"
            ),
            "classification": classify_progression(
                "AIM mean",
                aim_mean,
            ),
        }
    )

    major_percentage = major_n / total * 100
    major_ci = wilson_ci(major_n, total)

    rows.append(
        {
            "criterion": "Major technical issues",
            "numerator": major_n,
            "denominator": total,
            "observed_value": major_percentage,
            "unit": "percent of total sample",
            "ci_lower": major_ci[0] * 100,
            "ci_upper": major_ci[1] * 100,
            "protocol_thresholds": (
                "Go <10%; Amend 10-25%; Stop >25%"
            ),
            "classification": classify_progression(
                "Major technical issues",
                major_percentage,
            ),
        }
    )

    return pd.DataFrame(rows)


# =============================================================================
# MCQ item-level analysis
# =============================================================================

def parse_question_list(value: Any) -> set[int]:
    """Parse 'Q1, Q3, Q7' into {1, 3, 7}."""
    if pd.isna(value):
        return set()

    return {
        int(number)
        for number in re.findall(
            r"Q\s*([1-7])",
            str(value),
            flags=re.I,
        )
    }


def mcq_item_accuracy(
    df: pd.DataFrame,
    mcq_mask: pd.Series,
) -> pd.DataFrame:
    """
    Per-item correct percentages and exploratory exact McNemar tests.

    Among participants meeting the >=5/7 attempted threshold, absence from
    the correct-question list is treated as incorrect, consistent with the
    protocol's scoring rule for missing items.
    """
    eligible = df.loc[mcq_mask].copy()
    pre_sets = eligible[
        "pre_mcq_questions_correct"
    ].apply(parse_question_list)

    post_sets = eligible[
        "post_mcq_questions_correct"
    ].apply(parse_question_list)

    rows = []
    n = len(eligible)

    for question in range(1, N_MCQ_ITEMS + 1):
        pre_correct = pre_sets.apply(
            lambda values: question in values
        )

        post_correct = post_sets.apply(
            lambda values: question in values
        )

        both_incorrect = int(
            (~pre_correct & ~post_correct).sum()
        )

        incorrect_to_correct = int(
            (~pre_correct & post_correct).sum()
        )

        correct_to_incorrect = int(
            (pre_correct & ~post_correct).sum()
        )

        both_correct = int(
            (pre_correct & post_correct).sum()
        )

        discordant = (
            incorrect_to_correct
            + correct_to_incorrect
        )

        mcnemar_p = (
            float(
                binomtest(
                    incorrect_to_correct,
                    discordant,
                    p=0.5,
                    alternative="two-sided",
                ).pvalue
            )
            if discordant > 0
            else 1.0
        )

        rows.append(
            {
                "question": f"Q{question}",
                "paired_n": n,
                "pre_correct_n": int(
                    pre_correct.sum()
                ),
                "pre_correct_percentage": (
                    pre_correct.mean() * 100
                ),
                "post_correct_n": int(
                    post_correct.sum()
                ),
                "post_correct_percentage": (
                    post_correct.mean() * 100
                ),
                "percentage_point_change": (
                    (
                        post_correct.mean()
                        - pre_correct.mean()
                    )
                    * 100
                ),
                "both_incorrect_n": both_incorrect,
                "incorrect_to_correct_n": (
                    incorrect_to_correct
                ),
                "correct_to_incorrect_n": (
                    correct_to_incorrect
                ),
                "both_correct_n": both_correct,
                "exact_mcnemar_p": mcnemar_p,
                "holm_adjusted_p_across_7_items": np.nan,
            }
        )

    results = pd.DataFrame(rows)

    results[
        "holm_adjusted_p_across_7_items"
    ] = holm_adjust(
        results["exact_mcnemar_p"].tolist()
    )

    return results


# =============================================================================
# Qualitative-data preparation
# =============================================================================

def word_count(value: Any) -> int:
    """Simple word count for a free-text response."""
    if pd.isna(value):
        return 0

    text = str(value).strip()
    if not text:
        return 0

    return len(
        re.findall(
            r"\b\w+\b",
            text,
            flags=re.UNICODE,
        )
    )


def qualitative_summary(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """Response completeness and word-count descriptives."""
    rows = []

    for column, label in [
        (
            "most_useful_learning",
            "Most useful/memorable/surprising learning",
        ),
        (
            "improvements",
            "Disliked/confusing/improvement suggestions",
        ),
    ]:
        responses = (
            df[column]
            .dropna()
            .astype(str)
            .str.strip()
        )

        responses = responses[
            responses.ne("")
        ]

        counts = responses.apply(word_count)

        rows.append(
            {
                "question": label,
                "total_participants": len(df),
                "responses_n": len(responses),
                "response_percentage": (
                    len(responses) / len(df) * 100
                ),
                "median_words": (
                    float(counts.median())
                    if len(counts)
                    else np.nan
                ),
                "q1_words": (
                    float(counts.quantile(0.25))
                    if len(counts)
                    else np.nan
                ),
                "q3_words": (
                    float(counts.quantile(0.75))
                    if len(counts)
                    else np.nan
                ),
                "minimum_words": (
                    int(counts.min())
                    if len(counts)
                    else np.nan
                ),
                "maximum_words": (
                    int(counts.max())
                    if len(counts)
                    else np.nan
                ),
            }
        )

    return pd.DataFrame(rows)


def qualitative_template(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    De-linked template for manual descriptive thematic grouping.

    Original participant IDs are omitted. Free text must still be manually
    checked for identifying information before any sharing or quotation.
    """
    rows = []

    for row_number, (_, row) in enumerate(
        df.iterrows(),
        start=1,
    ):
        case_id = f"P{row_number:03d}"

        for column, question_label in [
            (
                "most_useful_learning",
                "Most useful/memorable/surprising learning",
            ),
            (
                "improvements",
                "Disliked/confusing/improvement suggestions",
            ),
        ]:
            response = row[column]

            if (
                pd.isna(response)
                or str(response).strip() == ""
            ):
                continue

            rows.append(
                {
                    "case_id": case_id,
                    "question": question_label,
                    "response_text": str(response).strip(),
                    "de_identification_review": "",
                    "initial_code_1": "",
                    "initial_code_2": "",
                    "descriptive_category_or_theme": "",
                    "coder_notes": "",
                    "second_reviewer_comments": "",
                    "final_consensus_code": "",
                }
            )

    return pd.DataFrame(rows)


# =============================================================================
# Reproducibility metadata and workbook formatting
# =============================================================================

def metadata() -> pd.DataFrame:
    """Record the main analysis decisions."""
    rows = [
        (
            "Study design",
            "Single-arm pre-post pilot feasibility study",
        ),
        (
            "Primary preliminary outcome",
            "Pre-post change in 7-item MCQ knowledge score",
        ),
        (
            "Secondary preliminary outcomes",
            "Pre-post change in MRI anxiety and willingness",
        ),
        (
            "Main analysis population",
            (
                "Protocol-consistent available-case sample "
                "defined separately for each outcome"
            ),
        ),
        (
            "MCQ usable-pair rule",
            (
                f">={MCQ_MIN_ATTEMPTED}/{N_MCQ_ITEMS} "
                "items attempted at both time points; "
                "missing items scored incorrect"
            ),
        ),
        (
            "Primary paired test",
            (
                "Two-sided exact conditional-permutation "
                "Wilcoxon signed-rank test; zero differences "
                "excluded from ranks"
            ),
        ),
        (
            "Sensitivity paired test",
            (
                "Two-sided exact sign test; zero/tied "
                "changes excluded"
            ),
        ),
        (
            "Standardised effect size",
            (
                "Matched-pairs rank-biserial correlation "
                "based on nonzero signed ranks"
            ),
        ),
        (
            "Unstandardised effect estimate",
            (
                "Hodges-Lehmann paired location shift "
                "(median of Walsh averages)"
            ),
        ),
        (
            "Confidence intervals",
            (
                f"{CONFIDENCE_LEVEL:.0%} percentile bootstrap "
                f"CIs using {BOOTSTRAP_REPLICATES:,} paired-case "
                f"resamples; seed {BOOTSTRAP_SEED}"
            ),
        ),
        (
            "Multiplicity",
            (
                "Primary MCQ p value unadjusted; "
                "anxiety and willingness adjusted together "
                "using Holm's procedure"
            ),
        ),
        (
            "Missing data",
            "Available-case analysis; no imputation",
        ),
        (
            "MCQ item analyses",
            (
                "Descriptive item accuracy plus exploratory "
                "exact McNemar tests with Holm adjustment "
                "across seven items"
            ),
        ),
        (
            "AIM scoring",
            (
                "Items scored 1-5; participant mean requires "
                f">={AIM_MIN_ITEMS_FOR_MEAN}/4 answered items"
            ),
        ),
        (
            "Qualitative analysis",
            (
                "Script creates a coding template; themes are "
                "not generated automatically"
            ),
        ),
        (
            "Meaningful-completer analysis",
            (
                "Sensitivity analysis only; main paired analyses "
                "are not restricted by meaningful completion"
            ),
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=["parameter", "value"],
    )


def software_versions() -> pd.DataFrame:
    """Software versions for reproducibility."""
    return pd.DataFrame(
        [
            ("Python", sys.version.replace("\n", " ")),
            ("Platform", platform.platform()),
            ("pandas", pd.__version__),
            ("numpy", np.__version__),
            ("scipy", scipy.__version__),
        ],
        columns=["software", "version"],
    )


def format_sheet(
    worksheet,
    dataframe: pd.DataFrame,
) -> None:
    """Apply readable formatting to an Excel worksheet."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for index, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        letter = get_column_letter(index)
        maximum_length = len(str(column))

        for value in dataframe[column].head(250):
            if pd.notna(value):
                maximum_length = max(
                    maximum_length,
                    len(str(value)),
                )

        if column in {
            "value",
            "protocol_thresholds",
            "scoring_rule",
            "issue",
            "observed_value",
        }:
            width = min(
                max(maximum_length + 2, 25),
                80,
            )
        else:
            width = min(
                max(maximum_length + 2, 10),
                32,
            )

        worksheet.column_dimensions[
            letter
        ].width = width

    worksheet.row_dimensions[1].height = 35


def write_results(
    sheets: dict[str, pd.DataFrame],
) -> None:
    """Write aggregate analysis tables."""
    RESULTS_WORKBOOK.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        RESULTS_WORKBOOK,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in sheets.items():
            safe_name = sheet_name[:31]
            dataframe.to_excel(
                writer,
                sheet_name=safe_name,
                index=False,
            )
            format_sheet(
                writer.sheets[safe_name],
                dataframe,
            )


def write_qualitative_workbook(
    summary: pd.DataFrame,
    coding_template: pd.DataFrame,
) -> None:
    """Write the separate free-text coding workbook."""
    readme = pd.DataFrame(
        {
            "instruction": [
                (
                    "This workbook supports manual descriptive "
                    "thematic grouping/content analysis."
                ),
                (
                    "Original participant IDs have been replaced "
                    "with sequential case IDs."
                ),
                (
                    "Manually review all free text for potentially "
                    "identifying information before sharing."
                ),
                (
                    "Do not publish this workbook without "
                    "de-identification and research-team approval."
                ),
                (
                    "Where feasible, have a second researcher "
                    "review coding and resolve differences by consensus."
                ),
            ]
        }
    )

    with pd.ExcelWriter(
        QUALITATIVE_WORKBOOK,
        engine="openpyxl",
    ) as writer:
        readme.to_excel(
            writer,
            sheet_name="README",
            index=False,
        )

        summary.to_excel(
            writer,
            sheet_name="response_summary",
            index=False,
        )

        coding_template.to_excel(
            writer,
            sheet_name="coding_template",
            index=False,
        )

        for name, dataframe in [
            ("README", readme),
            ("response_summary", summary),
            ("coding_template", coding_template),
        ]:
            format_sheet(
                writer.sheets[name],
                dataframe,
            )

        response_column = (
            coding_template.columns.get_loc(
                "response_text"
            )
            + 1
        )

        writer.sheets[
            "coding_template"
        ].column_dimensions[
            get_column_letter(response_column)
        ].width = 100


# =============================================================================
# Plain-text report
# =============================================================================

def build_report(
    df: pd.DataFrame,
    flow: pd.DataFrame,
    feasibility_table: pd.DataFrame,
    paired: pd.DataFrame,
    sensitivity: pd.DataFrame | None,
    aim_summary: pd.DataFrame,
    technical_summary: pd.DataFrame,
    telemetry_summary: pd.DataFrame,
) -> str:
    """Create a concise text summary for copying into ChatGPT."""
    lines = []

    lines.append("FINAL ANALYSIS SUMMARY")
    lines.append("=" * 78)
    lines.append(f"Input: {INPUT_FILE.resolve()}")
    lines.append(f"Participants: {len(df)}")
    lines.append("")

    lines.append("SAMPLE FLOW")
    lines.append("-" * 78)

    for _, row in flow.iterrows():
        lines.append(
            f"{row['sample_definition']}: "
            f"{int(row['n'])} "
            f"({row['percentage_of_total']:.1f}%)"
        )

    lines.append("")
    lines.append("PROTOCOL PROGRESSION CRITERIA")
    lines.append("-" * 78)

    for _, row in feasibility_table.iterrows():
        if "percent" in str(row["unit"]):
            observed = (
                f"{int(row['numerator'])}/"
                f"{int(row['denominator'])} "
                f"({row['observed_value']:.1f}%; "
                f"95% CI {row['ci_lower']:.1f}% "
                f"to {row['ci_upper']:.1f}%)"
            )
        elif row["criterion"] == "AIM mean":
            observed = (
                f"mean {row['observed_value']:.2f} "
                f"(95% CI {row['ci_lower']:.2f} "
                f"to {row['ci_upper']:.2f})"
            )
        else:
            observed = (
                f"{row['observed_value']:.0f} participants"
            )

        lines.append(
            f"{row['criterion']}: "
            f"{observed} -> "
            f"{row['classification']}"
        )

    lines.append("")
    lines.append("PAIRED PRE-POST OUTCOMES")
    lines.append("-" * 78)

    for _, row in paired.iterrows():
        lines.append(
            f"{row['outcome']} "
            f"({row['analysis_role']}; "
            f"n={int(row['paired_n'])})"
        )

        lines.append(
            f"  Pre median {row['pre_median']:.2f} "
            f"(IQR {row['pre_q1']:.2f}-"
            f"{row['pre_q3']:.2f}); "
            f"post median {row['post_median']:.2f} "
            f"(IQR {row['post_q1']:.2f}-"
            f"{row['post_q3']:.2f})"
        )

        lines.append(
            f"  Median change {row['change_median']:.2f} "
            f"(IQR {row['change_q1']:.2f}-"
            f"{row['change_q3']:.2f}); "
            f"Hodges-Lehmann shift "
            f"{row['hodges_lehmann_shift']:.2f} "
            f"(bootstrap 95% CI "
            f"{row['hodges_lehmann_ci_lower']:.2f} "
            f"to {row['hodges_lehmann_ci_upper']:.2f})"
        )

        lines.append(
            f"  Increased/unchanged/decreased: "
            f"{int(row['n_increased'])}/"
            f"{int(row['n_unchanged'])}/"
            f"{int(row['n_decreased'])}"
        )

        lines.append(
            f"  Exact Wilcoxon signed-rank: "
            f"W={row['wilcoxon_w_two_sided']:.1f}, "
            f"p={format_p(row['wilcoxon_exact_permutation_p'])}"
        )

        if row["analysis_role"] == "Secondary":
            lines.append(
                f"  Holm-adjusted p="
                f"{format_p(row['holm_adjusted_p_secondary'])}"
            )

        lines.append(
            f"  Rank-biserial r="
            f"{row['rank_biserial']:.3f} "
            f"(bootstrap 95% CI "
            f"{row['rank_biserial_ci_lower']:.3f} "
            f"to {row['rank_biserial_ci_upper']:.3f})"
        )

        lines.append(
            f"  Exact sign-test sensitivity p="
            f"{format_p(row['sign_test_exact_p'])}"
        )

    if sensitivity is not None:
        lines.append("")
        lines.append(
            "MEANINGFUL-COMPLETER SENSITIVITY"
        )
        lines.append("-" * 78)

        for _, row in sensitivity.iterrows():
            lines.append(
                f"{row['outcome']}: "
                f"n={int(row['paired_n'])}, "
                f"median change="
                f"{row['change_median']:.2f}, "
                f"Wilcoxon p="
                f"{format_p(row['wilcoxon_exact_permutation_p'])}, "
                f"rank-biserial r="
                f"{row['rank_biserial']:.3f}"
            )

    aim_row = aim_summary.iloc[0]

    lines.append("")
    lines.append("AIM ACCEPTABILITY")
    lines.append("-" * 78)
    lines.append(
        f"Participant AIM mean: "
        f"n={int(aim_row['participants_with_score'])}, "
        f"mean={aim_row['mean']:.2f} "
        f"(SD {aim_row['sd']:.2f}; "
        f"95% CI {aim_row['mean_ci_lower']:.2f} "
        f"to {aim_row['mean_ci_upper']:.2f}), "
        f"median={aim_row['median']:.2f} "
        f"(IQR {aim_row['q1']:.2f}-"
        f"{aim_row['q3']:.2f})"
    )

    lines.append(
        f"Exploratory Cronbach alpha="
        f"{aim_row['cronbach_alpha_complete_cases']:.3f} "
        f"(complete-case n="
        f"{int(aim_row['complete_case_n_for_alpha'])})"
    )

    lines.append("")
    lines.append("TECHNICAL ISSUES")
    lines.append("-" * 78)

    for _, row in technical_summary.iterrows():
        if row["count"] > 0:
            lines.append(
                f"{row['category']}: "
                f"{int(row['count'])} "
                f"({row['percentage_of_total_sample']:.1f}%)"
            )

    lines.append("")
    lines.append("TELEMETRY")
    lines.append("-" * 78)

    for _, row in telemetry_summary.iterrows():
        lines.append(
            f"{row['variable']}: "
            f"n={int(row['n'])}, "
            f"median={row['median']:.2f} "
            f"(IQR {row['q1']:.2f}-"
            f"{row['q3']:.2f}), "
            f"range {row['minimum']:.2f}-"
            f"{row['maximum']:.2f}"
        )

    gameplay = numeric(df["time_gameplay_sec"])

    lines.append(
        f"Gameplay <4 minutes: "
        f"{int((gameplay < SHORT_GAMEPLAY_SECONDS).sum())}; "
        f">40 minutes: "
        f"{int((gameplay > LONG_GAMEPLAY_SECONDS).sum())}"
    )

    lines.append("")
    lines.append("QUALITATIVE DATA")
    lines.append("-" * 78)
    lines.append(
        "A separate coding workbook was created. "
        "The script does not assign themes automatically; "
        "manual descriptive thematic grouping/content analysis "
        "and de-identification review are still required."
    )

    return "\n".join(lines)


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    df = load_data()
    validation_issues = validate_data(df)
    masks = build_masks(df)

    flow = sample_flow(df, masks)
    demographic_table = demographics(df)
    missingness_table = missingness(df)

    (
        aim_summary,
        aim_item_distributions,
        aim_score_distribution,
        aim_issues,
    ) = aim_outputs(df)

    technical_summary = technical_issues(df)
    telemetry_summary = telemetry(df)

    feasibility_table = feasibility(
        df=df,
        masks=masks,
        aim_summary=aim_summary,
        technical_summary=technical_summary,
    )

    paired_results = run_outcome_analyses(
        df=df,
        masks=masks,
        meaningful_only=False,
    )

    sensitivity_results = None

    if RUN_MEANINGFUL_COMPLETER_SENSITIVITY:
        sensitivity_results = run_outcome_analyses(
            df=df,
            masks=masks,
            meaningful_only=True,
        )

    mcq_items = mcq_item_accuracy(
        df,
        masks["mcq_usable"],
    )

    qualitative_descriptives = (
        qualitative_summary(df)
    )

    coding_template = qualitative_template(df)

    combined_issues = pd.concat(
        [
            validation_issues.assign(
                source="general_validation"
            ),
            aim_issues.assign(
                source="aim_scoring"
            ),
        ],
        ignore_index=True,
        sort=False,
    )

    sheets = {
        "analysis_metadata": metadata(),
        "software_versions": software_versions(),
        "sample_flow": flow,
        "demographics": demographic_table,
        "missingness": missingness_table,
        "feasibility": feasibility_table,
        "paired_outcomes": paired_results,
        "mcq_item_accuracy": mcq_items,
        "aim_summary": aim_summary,
        "aim_item_distributions": aim_item_distributions,
        "aim_score_distribution": aim_score_distribution,
        "technical_issues": technical_summary,
        "telemetry": telemetry_summary,
        "qualitative_summary": qualitative_descriptives,
        "validation_issues": combined_issues,
    }

    if sensitivity_results is not None:
        sheets["paired_sensitivity"] = (
            sensitivity_results
        )

    write_results(sheets)

    write_qualitative_workbook(
        summary=qualitative_descriptives,
        coding_template=coding_template,
    )

    report = build_report(
        df=df,
        flow=flow,
        feasibility_table=feasibility_table,
        paired=paired_results,
        sensitivity=sensitivity_results,
        aim_summary=aim_summary,
        technical_summary=technical_summary,
        telemetry_summary=telemetry_summary,
    )

    TEXT_REPORT.write_text(
        report,
        encoding="utf-8",
    )

    print(report)
    print("")
    print("=" * 78)
    print("FILES CREATED")
    print(
        f"Results workbook: "
        f"{RESULTS_WORKBOOK.resolve()}"
    )
    print(
        f"Text report: "
        f"{TEXT_REPORT.resolve()}"
    )
    print(
        f"Qualitative workbook: "
        f"{QUALITATIVE_WORKBOOK.resolve()}"
    )


if __name__ == "__main__":
    main()
